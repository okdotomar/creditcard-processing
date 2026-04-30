"""
build_report.py
Generates the transaction report Excel file from your DynamoDB CSV export.
Put your exported CSV in the same folder and run: python3 build_report.py
"""

import csv
import sys
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import PieChart, Reference
from openpyxl.chart.series import DataPoint

# ── Load CSV ──────────────────────────────────────────────────────────────────
CSV_FILE = "results.csv"  # rename to match your exported filename

try:
    with open(CSV_FILE, newline='', encoding='utf-8-sig') as f:
        rows = list(csv.DictReader(f))
except FileNotFoundError:
    print(f"ERROR: Could not find '{CSV_FILE}'")
    print("Rename your DynamoDB export to 'transaction_log.csv' and run again.")
    sys.exit(1)

print(f"Loaded {len(rows)} transactions from {CSV_FILE}")

# ── Parse data ────────────────────────────────────────────────────────────────
amounts      = []
merchants    = defaultdict(int)
banks        = defaultdict(int)
statuses     = defaultdict(int)

for row in rows:
    # Amount
    try:
        amt = float(row.get("amount", 0))
        amounts.append(amt)
    except (ValueError, TypeError):
        pass

    # Merchant
    m = row.get("merchant_name", "Unknown").strip()
    merchants[m] += 1

    # Bank
    b = row.get("bank_name", "Unknown").strip()
    banks[b] += 1

    # Status
    s = row.get("status", "Unknown").strip()
    statuses[s] += 1

total_txns   = len(rows)
max_amount   = max(amounts) if amounts else 0
avg_amount   = sum(amounts) / len(amounts) if amounts else 0
approved     = statuses.get("Approved", 0)
declined     = statuses.get("Declined", 0)
errors       = statuses.get("Error", 0)
failures     = declined + errors
success_pct  = (approved / total_txns * 100) if total_txns else 0
failure_pct  = (failures / total_txns * 100) if total_txns else 0

# ── Styles ────────────────────────────────────────────────────────────────────
NAVY    = "1F4E79"
BLUE    = "2E75B6"
LTBLUE  = "D6E4F0"
WHITE   = "FFFFFF"
GREEN   = "70AD47"
RED_C   = "FF4444"

def hfill(color): return PatternFill("solid", start_color=color)
def hfont(bold=False, color="000000", size=11):
    return Font(name="Arial", bold=bold, color=color, size=size)
def center(): return Alignment(horizontal="center", vertical="center")
def left():   return Alignment(horizontal="left",   vertical="center")
thin = Side(style="thin", color="BBBBBB")
def border(): return Border(left=thin, right=thin, top=thin, bottom=thin)

# ── Workbook ──────────────────────────────────────────────────────────────────
wb = Workbook()

# ══════════════════════════════════════════════════════════════════════════════
# Sheet 1 — Summary
# ══════════════════════════════════════════════════════════════════════════════
ws = wb.active
ws.title = "Summary"
ws.column_dimensions["A"].width = 32
ws.column_dimensions["B"].width = 22

# Title
ws.merge_cells("A1:B1")
ws["A1"] = "Transaction Log Report — Clearinghouse"
ws["A1"].font      = hfont(bold=True, color=WHITE, size=14)
ws["A1"].fill      = hfill(NAVY)
ws["A1"].alignment = center()
ws.row_dimensions[1].height = 32

# Section header
ws.merge_cells("A2:B2")
ws["A2"] = "Key Metrics"
ws["A2"].font      = hfont(bold=True, color=WHITE, size=11)
ws["A2"].fill      = hfill(BLUE)
ws["A2"].alignment = center()
ws.row_dimensions[2].height = 20

metrics = [
    ("Total Transactions",         total_txns,                   ""),
    ("Largest Transaction Amount", f"${max_amount:,.2f}",         ""),
    ("Average Transaction Amount", f"${avg_amount:,.2f}",         ""),
    ("Total Approved",             approved,                      ""),
    ("Total Declined / Errors",    failures,                      ""),
    ("Success Rate",               f"{success_pct:.1f}%",         ""),
    ("Failure Rate",               f"{failure_pct:.1f}%",         ""),
]

for i, (label, value, _) in enumerate(metrics):
    row  = i + 3
    fill = hfill(LTBLUE) if i % 2 == 0 else hfill(WHITE)
    ws[f"A{row}"] = label
    ws[f"B{row}"] = value
    ws[f"A{row}"].font      = hfont(bold=True, size=11)
    ws[f"B{row}"].font      = hfont(size=11)
    ws[f"A{row}"].fill      = fill
    ws[f"B{row}"].fill      = fill
    ws[f"A{row}"].alignment = left()
    ws[f"B{row}"].alignment = center()
    ws[f"A{row}"].border    = border()
    ws[f"B{row}"].border    = border()
    ws.row_dimensions[row].height = 22

# ══════════════════════════════════════════════════════════════════════════════
# Sheet 2 — Merchant Pie Chart
# ══════════════════════════════════════════════════════════════════════════════
wm = wb.create_sheet("Merchants")
wm.column_dimensions["A"].width = 28
wm.column_dimensions["B"].width = 16

wm.merge_cells("A1:B1")
wm["A1"] = "Transactions by Merchant"
wm["A1"].font      = hfont(bold=True, color=WHITE, size=13)
wm["A1"].fill      = hfill(NAVY)
wm["A1"].alignment = center()
wm.row_dimensions[1].height = 28

for col, label in zip(["A","B"], ["Merchant", "Transaction Count"]):
    wm[f"{col}2"] = label
    wm[f"{col}2"].font      = hfont(bold=True, color=WHITE)
    wm[f"{col}2"].fill      = hfill(BLUE)
    wm[f"{col}2"].alignment = center()
    wm[f"{col}2"].border    = border()
wm.row_dimensions[2].height = 20

for i, (name, count) in enumerate(sorted(merchants.items(), key=lambda x: -x[1])):
    row  = i + 3
    fill = hfill(LTBLUE) if i % 2 == 0 else hfill(WHITE)
    wm[f"A{row}"] = name
    wm[f"B{row}"] = count
    wm[f"A{row}"].font      = hfont(size=10)
    wm[f"B{row}"].font      = hfont(size=10)
    wm[f"A{row}"].fill      = fill
    wm[f"B{row}"].fill      = fill
    wm[f"A{row}"].alignment = left()
    wm[f"B{row}"].alignment = center()
    wm[f"A{row}"].border    = border()
    wm[f"B{row}"].border    = border()
    wm.row_dimensions[row].height = 18

# Pie chart — merchants
merchant_count = len(merchants)
chart_m = PieChart()
chart_m.title  = "Transactions by Merchant"
chart_m.style  = 10
chart_m.width  = 20
chart_m.height = 15

data_m   = Reference(wm, min_col=2, min_row=2, max_row=2 + merchant_count)
labels_m = Reference(wm, min_col=1, min_row=3, max_row=2 + merchant_count)
chart_m.add_data(data_m, titles_from_data=True)
chart_m.set_categories(labels_m)
chart_m.dataLabels            = chart_m.dataLabels or type('obj', (object,), {})()
wm.add_chart(chart_m, f"D3")

# ══════════════════════════════════════════════════════════════════════════════
# Sheet 3 — Bank Pie Chart
# ══════════════════════════════════════════════════════════════════════════════
wb2 = wb.create_sheet("Banks")
wb2.column_dimensions["A"].width = 28
wb2.column_dimensions["B"].width = 16

wb2.merge_cells("A1:B1")
wb2["A1"] = "Transactions by Bank"
wb2["A1"].font      = hfont(bold=True, color=WHITE, size=13)
wb2["A1"].fill      = hfill(NAVY)
wb2["A1"].alignment = center()
wb2.row_dimensions[1].height = 28

for col, label in zip(["A","B"], ["Bank", "Transaction Count"]):
    wb2[f"{col}2"] = label
    wb2[f"{col}2"].font      = hfont(bold=True, color=WHITE)
    wb2[f"{col}2"].fill      = hfill(BLUE)
    wb2[f"{col}2"].alignment = center()
    wb2[f"{col}2"].border    = border()
wb2.row_dimensions[2].height = 20

for i, (name, count) in enumerate(sorted(banks.items(), key=lambda x: -x[1])):
    row  = i + 3
    fill = hfill(LTBLUE) if i % 2 == 0 else hfill(WHITE)
    wb2[f"A{row}"] = name
    wb2[f"B{row}"] = count
    wb2[f"A{row}"].font      = hfont(size=10)
    wb2[f"B{row}"].font      = hfont(size=10)
    wb2[f"A{row}"].fill      = fill
    wb2[f"B{row}"].fill      = fill
    wb2[f"A{row}"].alignment = left()
    wb2[f"B{row}"].alignment = center()
    wb2[f"A{row}"].border    = border()
    wb2[f"B{row}"].border    = border()
    wb2.row_dimensions[row].height = 18

bank_count = len(banks)
chart_b = PieChart()
chart_b.title  = "Transactions by Bank"
chart_b.style  = 10
chart_b.width  = 20
chart_b.height = 15

data_b   = Reference(wb2, min_col=2, min_row=2, max_row=2 + bank_count)
labels_b = Reference(wb2, min_col=1, min_row=3, max_row=2 + bank_count)
chart_b.add_data(data_b, titles_from_data=True)
chart_b.set_categories(labels_b)
wb2.add_chart(chart_b, "D3")

# ── Save ──────────────────────────────────────────────────────────────────────
out = "transaction_report.xlsx"
wb.save(out)
print(f"\nReport saved to {out}")
print(f"\n── Key Answers ──────────────────────────────")
print(f"  Largest transaction:  ${max_amount:,.2f}")
print(f"  Average transaction:  ${avg_amount:,.2f}")
print(f"  Success rate:         {success_pct:.1f}%")
print(f"  Failure rate:         {failure_pct:.1f}%")